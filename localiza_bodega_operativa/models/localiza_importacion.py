import base64
from io import BytesIO
from odoo import api, fields, models, _
from odoo.exceptions import UserError

try:
    import openpyxl
except Exception:
    openpyxl = None

class LocalizaImportacion(models.Model):
    _name = 'localiza.importacion'
    _description = 'Importacion Excel bodega'
    _inherit = ['mail.thread', 'mail.activity.mixin']
    _order = 'create_date desc'

    name = fields.Char(string='Nombre', required=True, default='Importacion Excel')
    tipo = fields.Selection([
        ('auto','Detectar automaticamente'),
        ('insumos','Insumos / Uniformes'),
        ('gps','GPS'),
    ], default='auto', string='Tipo')
    archivo = fields.Binary(string='Archivo Excel', required=True)
    archivo_nombre = fields.Char(string='Nombre archivo')
    hoja = fields.Char(string='Hoja a importar')
    state = fields.Selection([('draft','Borrador'),('processed','Procesado'),('error','Error')], default='draft')
    line_ids = fields.One2many('localiza.importacion.linea', 'importacion_id', string='Lineas')
    total_lineas = fields.Integer(compute='_compute_total')
    errores = fields.Text(string='Errores / notas')

    def _compute_total(self):
        for rec in self:
            rec.total_lineas = len(rec.line_ids)

    def action_procesar(self):
        if not openpyxl:
            raise UserError(_('La libreria openpyxl no esta disponible en el servidor.'))
        for rec in self:
            rec.line_ids.unlink()
            data = base64.b64decode(rec.archivo)
            wb = openpyxl.load_workbook(BytesIO(data), read_only=True, data_only=True)
            ws = wb[rec.hoja] if rec.hoja and rec.hoja in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                raise UserError(_('El archivo no tiene datos.'))
            header = [str(h).strip() if h is not None else '' for h in rows[0]]
            upper = [h.upper() for h in header]
            def val(row, names):
                for n in names:
                    if n in upper:
                        i = upper.index(n)
                        return row[i] if i < len(row) else False
                return False
            created = 0
            gps_mode = rec.tipo == 'gps' or 'IMEI' in upper
            for row in rows[1:]:
                if not row or not any(row):
                    continue
                vals = {'importacion_id': rec.id, 'raw_data': str(row)[:1000]}
                if gps_mode:
                    vals.update({
                        'tipo': 'gps',
                        'imei': str(val(row, ['IMEI']) or '').strip(),
                        'modelo': str(val(row, ['MODELO']) or '').strip(),
                        'marca': str(val(row, ['MARCA']) or '').strip(),
                        'fecha_compra': val(row, ['FECHA DE COMPRA']) or False,
                        'proveedor': str(val(row, ['PROVEEDOR']) or '').strip(),
                        'factura': str(val(row, ['NO. FACTURA','NO FACTURA']) or '').strip(),
                        'serie': str(val(row, ['NO SERIE','NO. SERIE','SERIE']) or '').strip(),
                        'precio': float(val(row, ['PRECIO UNIDAD','COSTO']) or 0.0),
                        'iva': float(val(row, ['IVA']) or 0.0),
                        'placa': str(val(row, ['PLACA']) or '').strip(),
                        'cliente': str(val(row, ['CLIENTE']) or '').strip(),
                    })
                else:
                    vals.update({
                        'tipo': 'articulo',
                        'categoria_texto': str(val(row, ['CATEGORIA','CATEGORÍA']) or '').strip(),
                        'subcategoria': str(val(row, ['SUB CATEGORIA','SUB CATEGORÍA','SUBCATEGORIA']) or '').strip(),
                        'codigo': str(val(row, ['REFERENCIA INTERNA','CODIGO','CÓDIGO']) or '').strip(),
                        'nombre': str(val(row, ['NOMBRE']) or val(row, ['SUB CATEGORIA','SUB CATEGORÍA']) or '').strip(),
                        'talla': str(val(row, ['ATRIBUTOS DEL PRODUCTO / VALORES','TALLA']) or '').strip(),
                        'fecha_ingreso': val(row, ['FECHA DE INGRESO']) or False,
                        'serie': str(val(row, ['SERIE']) or '').strip(),
                        'factura': str(val(row, ['NO. FACTURA','NO FACTURA']) or '').strip(),
                        'costo': float(val(row, ['COSTO']) or 0.0),
                        'proveedor': str(val(row, ['PROVEEDOR']) or '').strip(),
                        'ubicacion_texto': str(val(row, ['UBICACION','UBICACIÓN']) or '').strip(),
                    })
                self.env['localiza.importacion.linea'].create(vals)
                created += 1
            rec.state = 'processed'
            rec.errores = 'Lineas leidas: %s desde hoja %s' % (created, ws.title)
        return True

    def action_crear_registros(self):
        for rec in self:
            for line in rec.line_ids:
                line.action_crear_registro()
        return True

class LocalizaImportacionLinea(models.Model):
    _name = 'localiza.importacion.linea'
    _description = 'Linea importada de Excel'
    _order = 'id'

    importacion_id = fields.Many2one('localiza.importacion', ondelete='cascade')
    tipo = fields.Selection([('articulo','Articulo'),('gps','GPS')], default='articulo')
    estado = fields.Selection([('pendiente','Pendiente'),('creado','Creado'),('error','Error')], default='pendiente')
    error = fields.Char(string='Error')
    categoria_texto = fields.Char(string='Categoria Excel')
    subcategoria = fields.Char(string='Subcategoria')
    codigo = fields.Char(string='Referencia interna')
    nombre = fields.Char(string='Nombre')
    talla = fields.Char(string='Talla')
    fecha_ingreso = fields.Date(string='Fecha ingreso')
    serie = fields.Char(string='Serie')
    factura = fields.Char(string='Factura')
    costo = fields.Float(string='Costo')
    proveedor = fields.Char(string='Proveedor')
    ubicacion_texto = fields.Char(string='Ubicacion')
    imei = fields.Char(string='IMEI')
    modelo = fields.Char(string='Modelo')
    marca = fields.Char(string='Marca')
    fecha_compra = fields.Date(string='Fecha compra')
    precio = fields.Float(string='Precio')
    iva = fields.Float(string='IVA')
    placa = fields.Char(string='Placa')
    cliente = fields.Char(string='Cliente')
    raw_data = fields.Text(string='Data original')

    def _categoria_articulo(self):
        text = ((self.categoria_texto or '') + ' ' + (self.subcategoria or '') + ' ' + (self.nombre or '')).lower()
        if 'uniform' in text or 'camisa' in text or 'pantal' in text or 'sueter' in text:
            return 'uniforme'
        if 'bota' in text:
            return 'bota'
        if 'gps' in text:
            return 'gps'
        return 'insumo'

    def action_crear_registro(self):
        for rec in self:
            try:
                if rec.tipo == 'gps':
                    if not rec.imei:
                        rec.write({'estado':'error','error':'Sin IMEI'})
                        continue
                    existing = self.env['localiza.gps'].search([('imei','=',rec.imei)], limit=1)
                    if not existing:
                        self.env['localiza.gps'].create({
                            'imei': rec.imei, 'modelo': rec.modelo, 'marca': rec.marca,
                            'fecha_compra': rec.fecha_compra, 'proveedor': rec.proveedor,
                            'factura': rec.factura, 'serie': rec.serie, 'precio': rec.precio,
                            'iva': rec.iva, 'placa': rec.placa, 'cliente': rec.cliente,
                        })
                    rec.write({'estado':'creado','error':False})
                else:
                    name = rec.nombre or rec.subcategoria or rec.codigo or 'Articulo importado'
                    existing = rec.codigo and self.env['localiza.articulo'].search([('codigo','=',rec.codigo)], limit=1)
                    if not existing:
                        self.env['localiza.articulo'].create({
                            'name': name, 'codigo': rec.codigo or False, 'categoria': rec._categoria_articulo(),
                            'subcategoria': rec.subcategoria, 'talla': rec.talla, 'serie': rec.serie,
                            'factura': rec.factura, 'fecha_ingreso': rec.fecha_ingreso,
                            'proveedor': rec.proveedor, 'costo': rec.costo,
                            'ubicacion_texto': rec.ubicacion_texto, 'cantidad': 1.0,
                        })
                    rec.write({'estado':'creado','error':False})
            except Exception as e:
                rec.write({'estado':'error','error':str(e)[:250]})
        return True
