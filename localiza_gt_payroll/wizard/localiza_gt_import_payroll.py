import base64
from io import BytesIO
from odoo import fields, models, _
from odoo.exceptions import UserError


class LocalizaGtImportPayroll(models.TransientModel):
    _name = 'localiza.gt.import.payroll.wizard'
    _description = 'Importar planilla Localiza desde Excel'

    file = fields.Binary(string='Archivo Excel .xlsx', required=True)
    filename = fields.Char(string='Nombre archivo')
    batch_id = fields.Many2one('localiza.gt.payroll.batch', string='Planilla destino', required=True)
    sheet_name = fields.Char(string='Hoja', help='Ejemplo: 1ra quincena, 2da quincena o Nomina Prestaciones. Si queda vacío usa la primera hoja.')

    def action_import(self):
        self.ensure_one()
        try:
            import openpyxl
        except Exception as exc:
            raise UserError(_('Debe estar instalada la librería openpyxl en Odoo.sh. Error: %s') % exc)
        data = base64.b64decode(self.file)
        wb = openpyxl.load_workbook(BytesIO(data), data_only=False)
        ws = wb[self.sheet_name] if self.sheet_name and self.sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        header_row = None
        headers = {}
        for row in range(1, min(ws.max_row, 15) + 1):
            values = [str(ws.cell(row, col).value or '').strip().lower() for col in range(1, ws.max_column + 1)]
            if 'empleado' in values and ('sueldo base' in values or 'sueldo' in ' '.join(values)):
                header_row = row
                for col, val in enumerate(values, start=1):
                    if val:
                        headers[val] = col
                break
        if not header_row:
            raise UserError(_('No encontré la fila de encabezados. Debe contener al menos Empleado y Sueldo Base.'))

        def col_value(row, *names):
            for name in names:
                key = name.lower()
                if key in headers:
                    return ws.cell(row, headers[key]).value
            return None

        created = 0
        Employee = self.env['hr.employee']
        for row in range(header_row + 1, ws.max_row + 1):
            employee_name = col_value(row, 'empleado', 'nombre')
            if not employee_name or str(employee_name).strip().lower().startswith('total'):
                continue
            if str(employee_name).strip().lower() in ('administración', 'operaciones', 'monitoreo', 'técnicos', 'tecnicos'):
                continue
            employee = Employee.search([('name', 'ilike', str(employee_name).strip())], limit=1)
            if not employee:
                employee = Employee.create({'name': str(employee_name).strip()})
            vals = {
                'batch_id': self.batch_id.id,
                'employee_id': employee.id,
                'department_name': col_value(row, 'departamento', 'puesto') or '',
                'job_name': col_value(row, 'puesto') or '',
                'days_worked': float(col_value(row, 'días laborados', 'dias laborados') or 0.0),
                'base_salary': float(col_value(row, 'sueldo base') or 0.0),
                'incentive_bonus': float(col_value(row, 'bonificación incentivo', 'bonificacion incentivo') or 0.0),
                'decree_bonus': float(col_value(row, 'bono dto', 'bono decreto') or 0.0),
                'first_half_amount': float(col_value(row, 'primera quincena') or 0.0),
                'isr': float(col_value(row, 'isr') or 0.0),
                'advances': float(col_value(row, 'anticipos') or 0.0),
                'other_deductions': float(col_value(row, 'otros descuentos') or 0.0),
            }
            self.env['localiza.gt.payroll.batch.line'].create(vals)
            created += 1
        self.batch_id.action_compute()
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {'title': _('Importación completada'), 'message': _('%s líneas importadas.') % created, 'type': 'success', 'sticky': False},
        }
